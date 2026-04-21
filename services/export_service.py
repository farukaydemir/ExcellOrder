import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Protection
from datetime import datetime
from models import Order

def export_order_to_excel(output_dir):
    order = Order.query.first()
    if not order or not order.items:
        return None
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Sipariş Formu"
    
    headers = ["SKU", "Ürün Adı", "Renk", "Birim Fiyat", "Adet", "Satır Toplamı"]
    ws.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = 15
    ws.column_dimensions['B'].width = 40 
        
    row_num = 2
    for item in order.items:
        prod = item.product
        ws.cell(row=row_num, column=1, value=prod.sku)
        ws.cell(row=row_num, column=2, value=prod.name)
        ws.cell(row=row_num, column=3, value=prod.color)
        
        price_cell = ws.cell(row=row_num, column=4, value=item.unit_price)
        price_cell.number_format = '#,##0.00 ₺'
        
        ws.cell(row=row_num, column=5, value=item.quantity)
        
        total_cell = ws.cell(row=row_num, column=6, value=f"=D{row_num}*E{row_num}")
        total_cell.number_format = '#,##0.00 ₺'
        row_num += 1
        
    ws.cell(row=row_num+1, column=5, value="Genel Toplam").font = Font(bold=True)
    grand_total_cell = ws.cell(row=row_num+1, column=6, value=f"=SUM(F2:F{row_num-1})")
    grand_total_cell.number_format = '#,##0.00 ₺'
    grand_total_cell.font = Font(bold=True)
    
    # Protecting the worksheet but keeping Adet unlocked
    ws.protection.enable()
    ws.protection.password = '12345'
    for r in range(2, row_num):
        qty_cell = ws.cell(row=r, column=5)
        qty_cell.protection = Protection(locked=False)
        qty_cell.fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid") # subtle highlight for editable cell
        qty_cell.alignment = Alignment(horizontal='center')
        
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"siparis_formu_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    wb.save(filepath)
    return filepath
